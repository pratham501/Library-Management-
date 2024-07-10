import openpyxl
import datetime
from prettytable import PrettyTable 
l=[]
dt=datetime.datetime.today().strftime(r'%d/%m/%Y')
def issue_book():                                     
    workbook=openpyxl.load_workbook('Library.xlsx')
    sheet=workbook['Books']
    print("Enter details for issuing a book")
    b_nm=input("Book name: ").strip()
    a_nm=input("Author name: ").strip()
    k=find_book(b_nm,a_nm)
    if k!=0:
        if sheet[f'C{k}'].value >0:
            g=True          
        else :
            g=False
    else:
        g=False
    if g:
        nm_red=input("Name of reader: ").strip()
        crd_num=input("Library id: ").strip()
        print("Date of issue: ",dt)
        m=sheet[f'C{k}'].value
        sheet[f'C{k}']=m-1
        print("Book issued\n")
        workbook.save('Library.xlsx')
    else :
        print("Book currently not available\n")

def return_book():                                     
    workbook=openpyxl.load_workbook('Library.xlsx')
    sheet=workbook['Books']
    print("Enter book details to return-")
    b_nm=input("Book name: ").strip()
    a_nm=input("Author name: ").strip()
    g=find_book(b_nm,a_nm)
    if g>0 :
        m=sheet[f'C{g}'].value
        if m<50:
            nm_red=input("Name of reader: ").strip()
            print("Date of return: ",dt)
            crd_num=input("Library id: ").strip()
            #print("m is :",m)
            sheet[f'C{g}']=m+1
            print("Book returned\n")
            workbook.save('Library.xlsx')
        else :
            print("50 books already in store\n")
    elif g==0 :
        print("Enter valid book\n")

def find_book(b_nm,a_nm):    # find_book cannot be used independetly, use search_book to search a book
    workbook=openpyxl.load_workbook('Library.xlsx')
    sheet=workbook['Books']
    for i in range(2,21):
        if (sheet[f'A{i}'].value) !=None:

            if b_nm.lower() == (sheet[f'A{i}'].value).lower() and a_nm.lower() == (sheet[f'B{i}'].value).lower() :
                k=True 
                break         
        else :
            k=False
    if k:
        return i
    else:
        return 0

def search_book():                                    
    workbook=openpyxl.load_workbook('Library.xlsx')
    sheet=workbook['Books']
    print("Enter details of book to be searched-")
    b_nm=input("Book Name: ").strip()
    a_nm=input("Author Name: ").strip()
    g=find_book(b_nm.lower(),a_nm.lower())
    if g==0 or int(sheet[f'C{g}'].value) == 0:
        print("Book not available\n")
    else :
        print("Book found\n")
    
def purchase_book():                                 
    workbook=openpyxl.load_workbook('Library.xlsx')
    sheet=workbook['Books']
    print("Enter details of book for purchasing-")
    b_nm=input("Book name: ").strip()
    a_nm=input("Author name: ").strip()
    g=find_book(b_nm,a_nm)
    if g>0 :
        m=sheet[f'C{g}'].value
        if m>0:
            while True:
                q=int(input("Quantity: "))
                if q<=m :
                    break
                else:
                    print(f"Only {m} books available in store\n")
            if q>0:
                sheet[f'C{g}']=m-q
                pr=sheet[f'D{g}'].value

                print("\nBook purchased\n")
                generate_bill(sheet[f'A{g}'].value,sheet[f'B{g}'].value,pr,q)
                workbook.save('Library.xlsx')
            else:
                print("Enter valid quantity\n")
        else :
            print("Book not available in store\n")
    elif g==0 :
        print("Enter valid book\n")

def generate_bill(bk_nm,a_nm,price,q):               # Cannot be used independently
    global dt
    amnt=price*q
    for i in range(42+len(bk_nm)):
        print("=",end="")
    print("\nDate-",dt)
    print(f"\n\t\tBook name: {bk_nm}\n\t\tAuthor name: {a_nm}\n\t\tPrice: ₹ {price}\n\t\tQuantity: {q}\n\t\tAmount: ₹ {float(amnt)}")
    print("")
    for i in range(42+len(bk_nm)):
        print("=",end="")

def add_libid():
    workbook=openpyxl.load_workbook('Library.xlsx')
    sheet=workbook['IDs']
    print("Enter details for new Libraray id-")
    name=input("Name: ").strip()
    lib_id=input("Library id: ").strip()
    if name=='' or lib_id=='':
        return print("Enter valid details\n")
    i=2
    while True:
        if sheet[f'A{i}'].value == None and sheet[f'B{i}'].value == None :
            break
        i += 1
    sheet[f'A{i}']=name

    sheet[f'B{i}']=lib_id
    print('ID added\n')
    workbook.save('Library.xlsx')

def remove_libid():
    workbook=openpyxl.load_workbook('Library.xlsx')
    sheet=workbook['IDs']
    print("Enter details to remove Libraray id-")
    name=input("Name: ").strip()
    lib_id=input("Library id: ").strip()
    i=2
    while True and sheet[f'A{i}'].value != None :
        if sheet[f'A{i}'].value == name and sheet[f'B{i}'].value == lib_id :
            sheet.delete_rows(i)
            print('ID removed\n')
            break
        i += 1
    else :
        print("No such ID found\n")
    workbook.save('Library.xlsx')

class book:
    
    def __init__(self,bk_nm,ath_nm,price,qn):
        self.__bk_nm=bk_nm
        self.__ath_nm=ath_nm
        self.__price=price
        self.__qn=qn
    t=0

    @staticmethod 
    def add_book():
        workbook=openpyxl.load_workbook('Library.xlsx')
        sheet=workbook['Books']
        ath_nm=''
        for j in range(2,21):
            if sheet[f'A{j}'].value==None:
                #print(f"No value found at A{j} cell")
                break
        k=j
        for a in range(j,21): 
            print("\nEnter details for new book to be added")
            bk_nm=input("Book name: ").strip()
            if bk_nm=='':
                break
            ath_nm=input("Author name: ").strip()
            if ath_nm=='':
                break
            g=find_book(bk_nm,ath_nm)
            if g>0:
                duplicate(g)
                return
            price=int(input("Price: ₹"))
            while True:
                st_qn=int(input("Quantity: ")) 
                if 50 > st_qn > 0 :
                    break
                else :
                    print("Storage capacity is 50, enter valid number") 
            c=a-k   # c is telling 0,1,2,3...
            bc=book(bk_nm,ath_nm,price,st_qn)
            b=0
            while b<3:

                i=input("Do you want to add another book? ")
                b += 1
                if i.upper().strip() in ("Y","YES"):
                    break
                elif i.upper().strip() in ("N","NO"):
                    break
                else:
                    print("Enter valid input")
            print(f"\n{c+1} book added")
            global l 
            l.append(bc)
            sheet[f"A{a}"]=l[c].__bk_nm
            sheet[f"B{a}"]=l[c].__ath_nm
            sheet[f"C{a}"]=l[c].__qn
            sheet[f"D{a}"]=float(l[c].__price)
            workbook.save('Library.xlsx')
            t=a
            if (i.upper()).strip() in ("N","NO") or b==3:
                break 
        if bk_nm=='' or ath_nm=='':
            print("\nEnter valid details\n")
        if bk_nm!='' and ath_nm!='':    
            book.show_books_added(l)
    @staticmethod
    def show_books_added(self):                       # Cannot be used independly, it is itself called in add_book(), it shows books added in add_book function
        global l
        book_table=PrettyTable(["Book Name","Author Name","Quantity","Price (in ₹)"])
        for i in range(0,len(l)):
            book_table.add_row([l[i].__bk_nm,l[i].__ath_nm,l[i].__qn,float(l[i].__price)])
        print("\nBooks that are added-\n")
        print(book_table)

def duplicate(g):                                     #Cannot be used independently
    workbook=openpyxl.load_workbook('Library.xlsx')
    sheet=workbook['Books']
    m=sheet[f'C{g}'].value
    if m==50:
        print("Maximum quantity is 50")
        print("50 books already in store")
        return
    print("Book already exist in store, adding to existing quantity")
    if g>0 :
        b=0
        while b<2:
            b += 1
            q=int(input("Quantity: "))
            if 0<q+m<=50:
                sheet[f'C{g}']=q+m
                print("Book added")
                break
            else:
                print("Storage capacity cannot exceed 50")
                print(f"Books already in storage are {m}, you can add {50-(m)} books")
        if b==3:
            print("Books not added ")
    workbook.save("Library.xlsx")

def create_xlfile():                                   # To be used once to create excel file
    workbook=openpyxl.Workbook('Library.xlsx')
    workbook.save('Library.xlsx')
    wb=openpyxl.load_workbook('Library.xlsx')
    sh1=wb.create_sheet('Books',0)
    sh2=wb.create_sheet("IDs",1)
    wb.save('Library.xlsx')
    sh1=wb['Books']
    sh2=wb['IDs']
    sh1.column_dimensions['A'].width=25
    sh1.column_dimensions['B'].width=15
    sh1['A1']='Book Name'
    sh1['B1']='Author Name'
    sh1['C1']='Quantity'
    sh1['D1']='Price(in ₹)'
    sh2.column_dimensions['A'].width=20
    sh2.column_dimensions['B'].width=15
    sh2['A1']="Name"
    sh2['B1']="Library IDs"
    wb.save('Library.xlsx')  
    print("\nExcel file created by name: Library\n")
    backup_xlfile()
    wb2=openpyxl.Workbook()
    sh2=wb2.active
    sh2["A99"]=1
    wb2.save('v2.xlsx')

def backup_xlfile():                                     # Librarian can backup we they think it is needed   
    wb1=openpyxl.load_workbook('Library.xlsx')
    wb2=openpyxl.Workbook('Backup_lib.xlsx')
    wb2.save('Backup_lib.xlsx')
    wb2=openpyxl.load_workbook('Backup_lib.xlsx')
    sh1=wb2.create_sheet('Books',0)
    sh0=wb2.create_sheet('IDs',1)
    sh1.column_dimensions['A'].width=25
    sh1.column_dimensions['B'].width=15
    sh1.column_dimensions['G'].width=35
    sh0.column_dimensions['A'].width=20
    sh0.column_dimensions['B'].width=15
    sh0.column_dimensions['F'].width=35
    sh0['F1']=f"Last backup on {dt}"
    sh1['G1']=f"Last backup on {dt}"
    sh2=wb1['Books']
    sh3=wb1['IDs']
    for i in range(1, sh2.max_row+1):
        for j in range(1, sh2.max_column+1):
            sh1.cell(row=i, column=j).value = sh2.cell(row=i, column=j).value
    for i in range(1, sh3.max_row+1):
        for j in range(1, sh3.max_column+1):
            sh0.cell(row=i, column=j).value = sh3.cell(row=i, column=j).value   
    wb1.save('Library.xlsx')
    wb2.save('Backup_lib.xlsx')
    print("\nBackup created\n")

a=0
try :
    wb=openpyxl.load_workbook('v2.xlsx')
    sh=wb.active
    a=sh['A99'].value
except :
    create_xlfile()
if a==None:
    print("a is", a)
    create_xlfile() 

# The functions below can be used when neccesary   
book.add_book()
search_book()
##backup_xlfile()
##add_libid()
issue_book()
##remove_libid()
return_book()
purchase_book()





            


